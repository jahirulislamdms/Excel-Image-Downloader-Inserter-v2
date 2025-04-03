import os
import requests
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from tkinter import filedialog, Tk, simpledialog
import time
from requests.exceptions import RequestException, ConnectionError

def download_image(url, save_path):
    while True:
        try:
            response = requests.get(url, timeout=10)
            if response.status_code == 200:
                with open(save_path, 'wb') as f:
                    f.write(response.content)
                return True
            else:
                print(f"Skipped (HTTP {response.status_code}): {url}")
                return False
        except ConnectionError:
            print("No internet connection. Retrying in 10 seconds...")
            time.sleep(10)
        except RequestException:
            print(f"Skipped (Invalid URL or other issue): {url}")
            return False

def process_excel():
    root = Tk()
    root.withdraw()

    # Select Excel file
    file_path = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel Files", "*.xlsx")])
    if not file_path:
        print("No file selected.")
        return

    # Load workbook
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Display column names with numbers
    column_names = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    column_options = "\n".join([f"{i + 1}: {col_name}" for i, col_name in enumerate(column_names)])

    # Ask user for image URL column
    col_number = simpledialog.askinteger("Select Image URL Column", f"Select the column number for image URLs:\n\n{column_options}\n\nEnter a number:")
    if not col_number or col_number < 1 or col_number > ws.max_column:
        print("Invalid image URL column selected.")
        return

    # Ask user for naming column
    name_col_number = simpledialog.askinteger("Select Naming Column", f"Select the column number to use as the filename:\n\n{column_options}\n\nEnter a number:")
    if not name_col_number or name_col_number < 1 or name_col_number > ws.max_column:
        print("Invalid naming column selected.")
        return

    # Insert new column for embedded images
    last_col = ws.max_column + 1
    new_col_letter = openpyxl.utils.get_column_letter(last_col)
    ws.cell(row=1, column=last_col, value="Downloaded Image")

    # Create directory for downloaded images
    image_dir = os.path.join(os.path.dirname(file_path), "downloaded_images")
    os.makedirs(image_dir, exist_ok=True)

    # Process each row
    for index, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        url = row[col_number - 1].value
        filename_value = row[name_col_number - 1].value

        if url and isinstance(url, str) and url.startswith("http") and filename_value:
            safe_filename = f"{str(filename_value)}.jpg"
            image_path = os.path.join(image_dir, safe_filename)

            success = download_image(url, image_path)
            if success:
                print(f"Downloaded & Saved: {safe_filename}")

                # Insert image into Excel
                img = ExcelImage(image_path)
                ws.add_image(img, f"{new_col_letter}{index}")

    # Save the updated file
    output_path = file_path.replace(".xlsx", "_with_images.xlsx")
    wb.save(output_path)
    print(f"Excel file saved: {output_path}")

if __name__ == "__main__":
    process_excel()
